from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

TOLERANCE = 1e-6
ITERATIONS = 25
OUTPUT_FILE = "metodos_numericos.xlsx"


def _style_header(ws, row, titles):
    for index, title in enumerate(titles, start=1):
        cell = ws.cell(row=row, column=index, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _add_bisection_sheet(wb):
    ws = wb.create_sheet("Biseccion")
    ws["A1"] = "Tolerancia"
    ws["B1"] = TOLERANCE

    header_row = 5
    _style_header(
        ws,
        header_row,
        ["Iter", "a", "b", "f(a)", "f(b)", "m", "f(m)", "Intervalo", "Error", "Condicion"],
    )
    start_row = header_row + 1

    ws.cell(row=start_row, column=1, value=0)
    ws.cell(row=start_row, column=2, value=-1.0)
    ws.cell(row=start_row, column=3, value=1.0)

    for row in range(start_row, start_row + ITERATIONS):
        if row > start_row:
            prev = row - 1
            ws.cell(row=row, column=1, value=f"=A{prev}+1")
            ws.cell(
                row=row,
                column=2,
                value=(
                    f"=IF(OR(ABS(G{prev})<=$B$1,I{prev}<=$B$1),"
                    f"B{prev},IF(D{prev}*G{prev}<0,B{prev},F{prev}))"
                ),
            )
            ws.cell(
                row=row,
                column=3,
                value=(
                    f"=IF(OR(ABS(G{prev})<=$B$1,I{prev}<=$B$1),"
                    f"C{prev},IF(D{prev}*G{prev}<0,F{prev},C{prev}))"
                ),
            )

        ws.cell(row=row, column=4, value=f"=EXP(B{row})-COS(B{row})")
        ws.cell(row=row, column=5, value=f"=EXP(C{row})-COS(C{row})")
        ws.cell(row=row, column=6, value=f"=(B{row}+C{row})/2")
        ws.cell(row=row, column=7, value=f"=EXP(F{row})-COS(F{row})")
        ws.cell(row=row, column=8, value=f"=C{row}-B{row}")
        ws.cell(row=row, column=9, value=f"=ABS(H{row})/2")
        ws.cell(row=row, column=10, value=f"=IF(I{row}<=$B$1,\"Cumple\",\"Continuar\")")

    ws.freeze_panes = f"A{start_row}"


def _add_newton_sheet(wb):
    ws = wb.create_sheet("Newton-Raphson")
    ws["A1"] = "Tolerancia"
    ws["B1"] = TOLERANCE

    header_row = 5
    _style_header(
        ws,
        header_row,
        ["Iter", "x_n", "f(x_n)", "f'(x_n)", "x_{n+1}", "Error", "Condicion"],
    )
    start_row = header_row + 1

    ws.cell(row=start_row, column=1, value=0)
    ws.cell(row=start_row, column=2, value=1.5)

    for row in range(start_row, start_row + ITERATIONS):
        if row > start_row:
            prev = row - 1
            ws.cell(row=row, column=1, value=f"=A{prev}+1")
            ws.cell(row=row, column=2, value=f"=E{prev}")

        ws.cell(
            row=row,
            column=3,
            value=f"=B{row}*ATAN(B{row}/2)+LN(B{row}^2+4)-3",
        )
        ws.cell(
            row=row,
            column=4,
            value=(
                f"=ATAN(B{row}/2)+B{row}/(2*(1+(B{row}/2)^2))"
                f"+2*B{row}/(B{row}^2+4)"
            ),
        )
        ws.cell(
            row=row,
            column=5,
            value=f"=IF(D{row}=0,B{row},B{row}-C{row}/D{row})",
        )
        ws.cell(row=row, column=6, value=f"=ABS(E{row}-B{row})")
        ws.cell(row=row, column=7, value=f"=IF(F{row}<=$B$1,\"Cumple\",\"Continuar\")")

    ws.freeze_panes = f"A{start_row}"


def _add_secant_sheet(wb):
    ws = wb.create_sheet("Secante")
    ws["A1"] = "Tolerancia"
    ws["B1"] = TOLERANCE

    header_row = 5
    _style_header(
        ws,
        header_row,
        ["Iter", "x_{n-1}", "x_n", "f(x_{n-1})", "f(x_n)", "x_{n+1}", "Error", "Condicion"],
    )
    start_row = header_row + 1

    ws.cell(row=start_row, column=1, value=0)
    ws.cell(row=start_row, column=2, value=0.5)
    ws.cell(row=start_row, column=3, value=1.0)

    for row in range(start_row, start_row + ITERATIONS):
        if row > start_row:
            prev = row - 1
            ws.cell(row=row, column=1, value=f"=A{prev}+1")
            ws.cell(row=row, column=2, value=f"=C{prev}")
            ws.cell(row=row, column=3, value=f"=F{prev}")

        ws.cell(
            row=row,
            column=4,
            value=(
                f"=0.5+0.25*B{row}^2-B{row}*SIN(B{row})"
                f"-0.5*COS(2*B{row})"
            ),
        )
        ws.cell(
            row=row,
            column=5,
            value=(
                f"=0.5+0.25*C{row}^2-C{row}*SIN(C{row})"
                f"-0.5*COS(2*C{row})"
            ),
        )
        ws.cell(
            row=row,
            column=6,
            value=(
                f"=IF(E{row}=D{row},C{row},"
                f"C{row}-E{row}*(C{row}-B{row})/(E{row}-D{row}))"
            ),
        )
        ws.cell(row=row, column=7, value=f"=ABS(F{row}-C{row})")
        ws.cell(row=row, column=8, value=f"=IF(G{row}<=$B$1,\"Cumple\",\"Continuar\")")

    ws.freeze_panes = f"A{start_row}"


def _add_fixed_point_sheet(wb):
    ws = wb.create_sheet("Punto Fijo")
    ws["A1"] = "Tolerancia"
    ws["B1"] = TOLERANCE

    header_row = 5
    _style_header(ws, header_row, ["Iter", "x_n", "g(x_n)", "Error", "Condicion"])
    start_row = header_row + 1

    ws.cell(row=start_row, column=1, value=0)
    ws.cell(row=start_row, column=2, value=1.0)

    for row in range(start_row, start_row + ITERATIONS):
        if row > start_row:
            prev = row - 1
            ws.cell(row=row, column=1, value=f"=A{prev}+1")
            ws.cell(row=row, column=2, value=f"=C{prev}")

        ws.cell(row=row, column=3, value=f"=SQRT((10-B{row}^3)/4)")
        ws.cell(row=row, column=4, value=f"=ABS(C{row}-B{row})")
        ws.cell(row=row, column=5, value=f"=IF(D{row}<=$B$1,\"Cumple\",\"Continuar\")")

    ws.freeze_panes = f"A{start_row}"


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    _add_bisection_sheet(wb)
    _add_newton_sheet(wb)
    _add_secant_sheet(wb)
    _add_fixed_point_sheet(wb)
    return wb


def main():
    wb = build_workbook()
    wb.save(OUTPUT_FILE)
    print(f"Archivo Excel generado: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
